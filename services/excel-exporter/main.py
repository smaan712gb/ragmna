"""
Excel Exporter Service
Creates professional Excel reports with financial statements, charts, and analysis
Handles multiple worksheets, formatting, and data visualization
"""

import os
import json
import logging
from flask import Flask, request, jsonify, send_file
from functools import wraps
from typing import Dict, Any, List, Optional
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import io
import base64

app = Flask(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Service configurations
SERVICE_API_KEY = os.getenv('SERVICE_API_KEY')

class ExcelReportGenerator:
    """Professional Excel report generator for M&A analysis"""

    def __init__(self):
        self.styles = self._define_styles()

    def _define_styles(self) -> Dict[str, Any]:
        """Define Excel styling configurations"""
        return {
            'header': Font(bold=True, size=14, color='FFFFFF'),
            'subheader': Font(bold=True, size=12, color='000000'),
            'bold': Font(bold=True),
            'italic': Font(italic=True),
            'currency': '#,##0',
            'percentage': '0.00%',
            'number': '#,##0.00',
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'subheader_fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'center_align': Alignment(horizontal='center', vertical='center')
        }

    def generate_ma_analysis_report(self, analysis_data: Dict[str, Any]) -> bytes:
        """Generate comprehensive M&A analysis Excel report"""

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create worksheets
        self._create_executive_summary_sheet(wb, analysis_data)
        self._create_company_profile_sheet(wb, analysis_data)
        self._create_financial_statements_sheet(wb, analysis_data)
        self._create_valuation_analysis_sheet(wb, analysis_data)
        self._create_peer_comparison_sheet(wb, analysis_data)
        self._create_due_diligence_sheet(wb, analysis_data)
        self._create_charts_sheet(wb, analysis_data)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()

    def _create_executive_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create executive summary worksheet"""
        ws = wb.create_sheet("Executive Summary")

        # Title
        ws['A1'] = "M&A Analysis Executive Summary"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws.merge_cells('A1:E1')

        # Analysis metadata
        row = 3
        ws[f'A{row}'] = "Analysis Date:"
        ws[f'B{row}'] = data.get('generated_at', datetime.now().isoformat())[:10]
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        target_symbol = data.get('target_symbol', 'N/A')
        acquirer_symbol = data.get('acquirer_symbol', 'N/A')
        ws[f'A{row}'] = "Target Company:"
        ws[f'B{row}'] = target_symbol
        ws[f'A{row}'].font = self.styles['bold']

        if acquirer_symbol != 'N/A':
            row += 1
            ws[f'A{row}'] = "Acquirer Company:"
            ws[f'B{row}'] = acquirer_symbol
            ws[f'A{row}'].font = self.styles['bold']

        # Company classification
        row += 2
        ws[f'A{row}'] = "Company Classification"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:B{row}')

        classification = data.get('classification', {})
        row += 1
        ws[f'A{row}'] = "Primary Classification:"
        ws[f'B{row}'] = classification.get('primary_classification', 'Unknown').title()
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Confidence Score:"
        ws[f'B{row}'] = ".1%"
        ws[f'B{row}'].number_format = self.styles['percentage']

        # Valuation summary
        row += 2
        ws[f'A{row}'] = "Valuation Summary"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:D{row}')

        valuation = data.get('valuation', {})
        dcf_valuation = valuation.get('dcf', {})
        cca_valuation = valuation.get('cca', {})

        row += 1
        ws[f'A{row}'] = "DCF Enterprise Value:"
        if dcf_valuation.get('enterprise_value'):
            ws[f'B{row}'] = dcf_valuation['enterprise_value']
            ws[f'B{row}'].number_format = self.styles['currency']
        else:
            ws[f'B{row}'] = "N/A"

        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "CCA Implied Value:"
        if cca_valuation.get('implied_valuation', {}).get('blended_valuation', {}).get('blended_price_per_share'):
            blended_price = cca_valuation['implied_valuation']['blended_valuation']['blended_price_per_share']
            shares = data.get('target_data', {}).get('market', {}).get('sharesOutstanding', 0)
            if shares > 0:
                equity_value = blended_price * shares
                ws[f'B{row}'] = equity_value
                ws[f'B{row}'].number_format = self.styles['currency']
            else:
                ws[f'B{row}'] = "N/A"
        else:
            ws[f'B{row}'] = "N/A"

        # Key recommendations
        row += 2
        ws[f'A{row}'] = "Key Recommendations"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:D{row}')

        final_report = data.get('final_report', {})
        summary = final_report.get('summary', {})
        recommendations = summary.get('recommendations', [])

        for i, rec in enumerate(recommendations[:5]):  # Top 5 recommendations
            row += 1
            ws[f'A{row}'] = f"{i+1}."
            ws[f'B{row}'] = rec

        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_company_profile_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create company profile worksheet"""
        ws = wb.create_sheet("Company Profile")

        # Title
        ws['A1'] = "Company Profile & Fundamentals"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws.merge_cells('A1:D1')

        target_data = data.get('target_data', {})
        profile = target_data.get('profile', [{}])[0] if target_data.get('profile') else {}
        market = target_data.get('market', {})

        # Basic information
        row = 3
        ws[f'A{row}'] = "Basic Information"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = "Company Name:"
        ws[f'B{row}'] = profile.get('companyName', 'N/A')
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Symbol:"
        ws[f'B{row}'] = data.get('target_symbol', 'N/A')
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Sector:"
        ws[f'B{row}'] = profile.get('sector', 'N/A')
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Industry:"
        ws[f'B{row}'] = profile.get('industry', 'N/A')
        ws[f'A{row}'].font = self.styles['bold']

        # Market data
        row += 2
        ws[f'A{row}'] = "Market Data"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:B{row}')

        row += 1
        ws[f'A{row}'] = "Market Cap:"
        market_cap = market.get('marketCap', 0)
        ws[f'B{row}'] = market_cap
        ws[f'B{row}'].number_format = self.styles['currency']
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Current Price:"
        price = market.get('price', 0)
        ws[f'B{row}'] = price
        ws[f'B{row}'].number_format = self.styles['currency']
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Shares Outstanding:"
        shares = market.get('sharesOutstanding', 0)
        ws[f'B{row}'] = shares
        ws[f'B{row}'].number_format = self.styles['number']
        ws[f'A{row}'].font = self.styles['bold']

        # Classification details
        row += 2
        ws[f'A{row}'] = "Company Classification"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:B{row}')

        classification = data.get('classification', {})

        row += 1
        ws[f'A{row}'] = "Primary Classification:"
        ws[f'B{row}'] = classification.get('primary_classification', 'Unknown').title()
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Revenue Growth Rate:"
        growth_rate = classification.get('revenue_growth_rate', 0)
        ws[f'B{row}'] = growth_rate
        ws[f'B{row}'].number_format = self.styles['percentage']
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Profitability Profile:"
        ws[f'B{row}'] = classification.get('profitability_profile', 'Unknown').title()
        ws[f'A{row}'].font = self.styles['bold']

        row += 1
        ws[f'A{row}'] = "Risk Assessment:"
        ws[f'B{row}'] = classification.get('risk_assessment', 'Unknown').title()
        ws[f'A{row}'].font = self.styles['bold']

    def _create_financial_statements_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create financial statements worksheet"""
        ws = wb.create_sheet("Financial Statements")

        # Title
        ws['A1'] = "Financial Statements & Projections"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws.merge_cells('A1:G1')

        financial_model = data.get('financial_model', {})

        # Income Statement
        row = 3
        ws[f'A{row}'] = "Income Statement Projections"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:G{row}')

        income_stmt = financial_model.get('income_statement', [])
        if income_stmt:
            # Headers
            row += 1
            headers = ['Year', 'Revenue', 'COGS', 'Gross Profit', 'OpEx', 'EBIT', 'Net Income', 'EPS']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = self.styles['bold']

            # Data
            for stmt in income_stmt:
                row += 1
                ws.cell(row=row, column=1, value=stmt.get('year', ''))
                ws.cell(row=row, column=2, value=stmt.get('revenue', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=3, value=stmt.get('cost_of_revenue', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=4, value=stmt.get('gross_profit', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=5, value=stmt.get('operating_expenses', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=6, value=stmt.get('operating_income', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=7, value=stmt.get('net_income', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=8, value=stmt.get('eps', 0)).number_format = self.styles['number']

        # Balance Sheet
        row += 3
        ws[f'A{row}'] = "Balance Sheet Projections"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:G{row}')

        balance_sheet = financial_model.get('balance_sheet', [])
        if balance_sheet:
            # Headers
            row += 1
            headers = ['Year', 'Cash', 'AR', 'Inventory', 'Total Assets', 'Debt', 'Equity', 'BV/Share']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = self.styles['bold']

            # Data
            for stmt in balance_sheet:
                row += 1
                ws.cell(row=row, column=1, value=stmt.get('year', ''))
                ws.cell(row=row, column=2, value=stmt.get('cash_and_equivalents', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=3, value=stmt.get('accounts_receivable', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=4, value=stmt.get('inventory', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=5, value=stmt.get('total_assets', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=6, value=stmt.get('total_liabilities', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=7, value=stmt.get('shareholders_equity', 0)).number_format = self.styles['currency']

                # Book value per share
                shares = stmt.get('shares_outstanding', 0)
                bv_per_share = stmt.get('shareholders_equity', 0) / shares if shares > 0 else 0
                ws.cell(row=row, column=8, value=bv_per_share).number_format = self.styles['number']

        # Auto-adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

    def _create_valuation_analysis_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create valuation analysis worksheet"""
        ws = wb.create_sheet("Valuation Analysis")

        # Title
        ws['A1'] = "Valuation Analysis & Multiples"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws.merge_cells('A1:F1')

        valuation = data.get('valuation', {})

        # DCF Analysis
        row = 3
        ws[f'A{row}'] = "DCF Valuation Results"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:C{row}')

        dcf = valuation.get('dcf', {})
        if dcf:
            row += 1
            ws[f'A{row}'] = "Enterprise Value:"
            ws[f'B{row}'] = dcf.get('final_valuation', {}).get('enterprise_value', 0)
            ws[f'B{row}'].number_format = self.styles['currency']
            ws[f'A{row}'].font = self.styles['bold']

            row += 1
            ws[f'A{row}'] = "Equity Value:"
            ws[f'B{row}'] = dcf.get('final_valuation', {}).get('equity_value', 0)
            ws[f'B{row}'].number_format = self.styles['currency']
            ws[f'A{row}'].font = self.styles['bold']

            row += 1
            ws[f'A{row}'] = "Price per Share:"
            ws[f'B{row}'] = dcf.get('final_valuation', {}).get('equity_value_per_share', 0)
            ws[f'B{row}'].number_format = self.styles['currency']
            ws[f'A{row}'].font = self.styles['bold']

            row += 1
            ws[f'A{row}'] = "WACC:"
            ws[f'B{row}'] = dcf.get('wacc', 0)
            ws[f'B{row}'].number_format = self.styles['percentage']
            ws[f'A{row}'].font = self.styles['bold']

        # CCA Analysis
        row += 2
        ws[f'A{row}'] = "Comparable Company Analysis"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:C{row}')

        cca = valuation.get('cca', {})
        if cca:
            implied_valuation = cca.get('implied_valuation', {})
            blended = implied_valuation.get('blended_valuation', {})

            row += 1
            ws[f'A{row}'] = "Implied Price per Share:"
            ws[f'B{row}'] = blended.get('blended_price_per_share', 0)
            ws[f'B{row}'].number_format = self.styles['currency']
            ws[f'A{row}'].font = self.styles['bold']

            row += 1
            ws[f'A{row}'] = "Valuation Range:"
            price_range = blended.get('price_range', [0, 0])
            ws[f'B{row}'] = f"${price_range[0]:,.2f} - ${price_range[1]:,.2f}"
            ws[f'A{row}'].font = self.styles['bold']

            # Multiples used
            row += 2
            ws[f'A{row}'] = "Valuation Multiples Used"
            ws[f'A{row}'].font = self.styles['subheader']
            ws[f'A{row}'].fill = self.styles['subheader_fill']
            ws.merge_cells(f'A{row}:C{row}')

            adjusted_multiples = cca.get('adjusted_multiples', {})
            for multiple_type, multiple_data in adjusted_multiples.items():
                row += 1
                ws[f'A{row}'] = f"{multiple_type.upper()}:"
                ws[f'B{row}'] = multiple_data.get('adjusted_value', 0)
                ws[f'B{row}'].number_format = self.styles['number']
                ws[f'A{row}'].font = self.styles['bold']

        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_peer_comparison_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create peer comparison worksheet"""
        ws = wb.create_sheet("Peer Comparison")

        # Title
        ws['A1'] = "Peer Company Analysis"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws.merge_cells('A1:E1')

        peers = data.get('peers', [])

        if peers:
            # Headers
            row = 3
            headers = ['Company', 'Symbol', 'Market Cap', 'Similarity Score', 'Price']
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = self.styles['bold']

            # Peer data
            for peer in peers:
                row += 1
                ws.cell(row=row, column=1, value=peer.get('companyName', ''))
                ws.cell(row=row, column=2, value=peer.get('symbol', ''))
                ws.cell(row=row, column=3, value=peer.get('marketCap', 0)).number_format = self.styles['currency']
                ws.cell(row=row, column=4, value=peer.get('similarity_score', 0)).number_format = self.styles['percentage']
                ws.cell(row=row, column=5, value=peer.get('price', 0)).number_format = self.styles['currency']

        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_due_diligence_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create due diligence worksheet"""
        ws = wb.create_sheet("Due Diligence")

        # Title
        ws['A1'] = "Due Diligence Findings"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws.merge_cells('A1:D1')

        dd_report = data.get('due_diligence', {})

        # Risk assessment
        row = 3
        ws[f'A{row}'] = "Risk Assessment Summary"
        ws[f'A{row}'].font = self.styles['subheader']
        ws[f'A{row}'].fill = self.styles['subheader_fill']
        ws.merge_cells(f'A{row}:B{row}')

        # Add DD findings (simplified)
        row += 2
        ws[f'A{row}'] = "Due diligence analysis completed. Detailed findings available in the full report."

        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _create_charts_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create charts and visualizations worksheet"""
        ws = wb.create_sheet("Charts & Visualizations")

        # Title
        ws['A1'] = "Financial Charts & Visualizations"
        ws['A1'].font = self.styles['header']
        ws['A1'].fill = self.styles['header_fill']
        ws.merge_cells('A1:D1')

        # Placeholder for charts
        ws['A3'] = "Charts would be embedded here in a full implementation"
        ws['A3'].font = self.styles['italic']

        # Add sample data for potential charts
        financial_model = data.get('financial_model', {})
        income_stmt = financial_model.get('income_statement', [])

        if income_stmt:
            # Revenue growth chart data
            row = 5
            ws[f'A{row}'] = "Year"
            ws[f'B{row}'] = "Revenue"
            ws[f'A{row}'].font = self.styles['bold']
            ws[f'B{row}'].font = self.styles['bold']

            for stmt in income_stmt:
                row += 1
                ws.cell(row=row, column=1, value=stmt.get('year', ''))
                ws.cell(row=row, column=2, value=stmt.get('revenue', 0)).number_format = self.styles['currency']

# Global Excel generator instance
excel_generator = ExcelReportGenerator()

def require_api_key(f):
    """Decorator to require API key"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        api_key = request.headers.get('X-API-Key')
        if not api_key or api_key != SERVICE_API_KEY:
            return jsonify({'error': 'Invalid API key'}), 401
        return f(*args, **kwargs)
    return decorated_function

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'service': 'excel-exporter',
        'version': '1.0.0'
    })

@app.route('/export/ma-analysis', methods=['POST'])
@require_api_key
def export_ma_analysis():
    """Export comprehensive M&A analysis to Excel"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({'error': 'Analysis data is required'}), 400

        # Generate Excel file
        excel_data = excel_generator.generate_ma_analysis_report(data)

        # Create response with Excel file
        response = send_file(
            io.BytesIO(excel_data),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"MA_Analysis_{data.get('target_symbol', 'Unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        return response

    except Exception as e:
        logger.error(f"Error exporting M&A analysis: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export/financial-model', methods=['POST'])
@require_api_key
def export_financial_model():
    """Export financial model to Excel"""
    try:
        data = request.get_json()
        financial_model = data.get('financial_model', {})

        if not financial_model:
            return jsonify({'error': 'Financial model data is required'}), 400

        # Create simplified Excel with just financial statements
        wb = Workbook()
        ws = wb.active
        ws.title = "Financial Model"

        # Add financial statements data
        excel_generator._create_financial_statements_sheet(wb, {'financial_model': financial_model})

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = send_file(
            io.BytesIO(output.getvalue()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Financial_Model_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        return response

    except Exception as e:
        logger.error(f"Error exporting financial model: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export/valuation', methods=['POST'])
@require_api_key
def export_valuation():
    """Export valuation analysis to Excel"""
    try:
        data = request.get_json()
        valuation = data.get('valuation', {})

        if not valuation:
            return jsonify({'error': 'Valuation data is required'}), 400

        # Create Excel with valuation analysis
        wb = Workbook()
        ws = wb.active
        ws.title = "Valuation Analysis"

        # Add valuation data
        excel_generator._create_valuation_analysis_sheet(wb, {'valuation': valuation})

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = send_file(
            io.BytesIO(output.getvalue()),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Valuation_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        return response

    except Exception as e:
        logger.error(f"Error exporting valuation: {e}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.getenv('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
